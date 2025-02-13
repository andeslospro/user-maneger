from tkinter import ttk
import customtkinter
from openpyxl import Workbook, load_workbook
from CTkMessagebox import CTkMessagebox
import openpyxl
import sqlite3
import requests
class MainApp(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Maneger")
        self.geometry("700x450")
        self.minsize(700, 450)
        self.maxsize(700, 450)

        #---------------------Configure grid layout---------------------------------------------------
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        #---------------------Creating Menu Frame----------------------------------------------------
        self.menu_frame = customtkinter.CTkFrame(self)
        self.menu_frame.grid(row=0, column=0, sticky="ns")

        #----------------------Creating Menu Buttons-------------------------------------------------
        self.home_button = customtkinter.CTkButton(self.menu_frame, fg_color="transparent", text="Home", command=self.home_button_event)
        self.home_button.pack(pady=10, fill="x")

        self.utilitys_button = customtkinter.CTkButton(self.menu_frame, fg_color="transparent", text="Utilidades", command=self.utilitys_button_event)
        self.utilitys_button.pack(pady=10, fill="x")

        self.members_button = customtkinter.CTkButton(self.menu_frame, fg_color="transparent", text="Miembros", command=self.members_button_event)
        self.members_button.pack(pady=10, fill="x")
        #----------------Login Frame---------------------------------------------------------
        self.login_frame = customtkinter.CTkFrame(self, fg_color="light gray")
        self.login_frame.grid(row=0, column=0, columnspan=2, sticky="nswe")

        self.login_label = customtkinter.CTkLabel(self.login_frame, text="Login", text_color="black",
                                                  font=("Arial", 24))
        self.login_label.pack(pady=20)

        self.lo_username_entry = customtkinter.CTkEntry(self.login_frame, placeholder_text="Username")
        self.lo_username_entry.pack(pady=10, padx=20, fill="x")

        self.lo_password_entry = customtkinter.CTkEntry(self.login_frame, placeholder_text="Password", show="*")
        self.lo_password_entry.pack(pady=10, padx=20, fill="x")

        self.lo_login_button = customtkinter.CTkButton(self.login_frame, text="Login", command=self.check_login)
        self.lo_login_button.pack(pady=15)

        self.lo_registe_button = customtkinter.CTkButton(master=self.login_frame, text="Register", command=self.register_window)
        self.lo_registe_button.pack(pady=15)

        #----------------Register Frame -------------------------------------------------
        self.register_frame = customtkinter.CTkFrame(self, fg_color="light gray")
        self.register_frame.grid(row=0, column=0, columnspan=2, sticky="nswe")

        self.login_label = customtkinter.CTkLabel(self.register_frame, text="Register", text_color="black",
                                                  font=("Arial", 24))
        self.login_label.pack(pady=20)

        self.re_username_entry = customtkinter.CTkEntry(self.register_frame, placeholder_text="Username")
        self.re_username_entry.pack(pady=10, padx=20, fill="x")

        self.re_password_entry = customtkinter.CTkEntry(self.register_frame, placeholder_text="Password", show="*")
        self.re_password_entry.pack(pady=10, padx=20, fill="x")

        self.re_register_but= customtkinter.CTkButton(self.register_frame, text="Register", command=self.check_register)
        self.re_register_but.pack(pady=15)
        
        self.homebut = customtkinter.CTkButton(self.register_frame, text="Home", command=self.login_frame.tkraise)
        self.homebut.pack(pady=15)

    
        #----------------Home Frame---------------------------------------------------
        self.home_frame = customtkinter.CTkFrame(self, fg_color="light gray")
        self.home_frame.grid(row=0, column=1, sticky="nswe")

        self.home_label = customtkinter.CTkLabel(self.home_frame, text="Home", text_color="black",font=("Arial", 24))
        self.home_label.pack(pady=20)

        #------------------Utilitys Frame---------------------------------------------
        self.utilitys_frame = customtkinter.CTkFrame(self, fg_color="light gray")
        self.utilitys_frame.grid(row=0, column=1, sticky="nswe")

        self.utilitys_label = customtkinter.CTkLabel(self.utilitys_frame, text="Utilitys", text_color="black", font=("Arial", 24))
        self.utilitys_label.pack(pady=20)

        #------------------Members Frame----------------------------------------------
        self.members_frame = customtkinter.CTkFrame(self, fg_color="light gray")
        self.members_frame.grid(row=0, column=1, sticky="nswe")

        # Configure members_frame to expand
        self.members_frame.grid_rowconfigure(0, weight=1)
        self.members_frame.grid_columnconfigure(0, weight=1)

        #-----------------Two main frames inside Members_Frame------------------------
        self.members_frame2 = customtkinter.CTkFrame(self.members_frame, fg_color="blue")
        self.members_frame2.grid(row=0, column=0, sticky="nswe")

        self.members_frame3 = customtkinter.CTkFrame(self.members_frame, fg_color="blue")
        self.members_frame3.grid(row=1, column=0, sticky="nswe")

        self.members_frame2.grid_rowconfigure(0, weight=1)
        self.members_frame2.grid_columnconfigure(0, weight=1)
        self.members_frame2.grid_columnconfigure(1, weight=1)
        
        self.members_frame3.grid_rowconfigure(0, weight=1)
        self.members_frame3.grid_columnconfigure(0, weight=1)

        self.frame_1 = customtkinter.CTkFrame(self.members_frame2)
        self.frame_1.grid(row=0, column=0, sticky="nswe")

        self.frame_2 = customtkinter.CTkFrame(self.members_frame2)
        self.frame_2.grid(row=0, column=1, sticky="nswe")

        self.label1 = customtkinter.CTkLabel(self.frame_1, text="Nombre", text_color="black")
        self.label1.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.name1 = customtkinter.CTkEntry(self.frame_1, placeholder_text="Nombre")
        self.name1.grid(row=0, column=1, padx=5, pady=5, sticky="we")

        self.label2 = customtkinter.CTkLabel(self.frame_1, text="Apellido", text_color="black")
        self.label2.grid(row=1, column=0, padx=5, pady=5, sticky="w")

        self.apellido1 = customtkinter.CTkEntry(self.frame_1, placeholder_text="Apellido")
        self.apellido1.grid(row=1, column=1, padx=5, pady=5, sticky="we")

        self.submit_button = customtkinter.CTkButton(self.frame_1, command=self.getvalue, text="Save In Excel")
        self.submit_button.grid(row=2, column=1, padx=5, pady=5, sticky="we")

        self.submit_button2 = customtkinter.CTkButton(self.frame_1, command=self.load_data, text="LoadData", state="normal")
        self.submit_button2.grid(row=3, column=1, padx=5, pady=5, sticky="we")

        self.cedula = customtkinter.CTkEntry(master=self.frame_1, placeholder_text="ID")
        self.cedula.grid(row=0, column=2, pady=5, padx=5, sticky="we")
        
        self.edad = customtkinter.CTkEntry(master=self.frame_1, placeholder_text="Edad")
        self.edad.grid(row=1, column=2)
        self.online_off = customtkinter.CTkSwitch(master=self.frame_1)
        self.online_off.grid(row=2, column=2)

        # Show XLSX Files
        self.treeview_columns = ("Nombre", "Apellido", "Edad", "ID")
        self.treevie = ttk.Treeview(self.members_frame3, show="headings", columns=self.treeview_columns, height=20)
        self.treevie.grid(row=0, column=0, sticky="nswe")

        

        self.treevie.column("Nombre", width=100) 
        self.treevie.column("Apellido", width=100, minwidth=50)
        self.treevie.column("Edad", width=50)
        self.treevie.column("ID", width=100)

        for col in self.treeview_columns:
            self.treevie.heading(col, text=col)

        
        #-------------Default Frame----------------------------------
        self.select_frame_by_name("login")

    # ----------------Selected Frame---------------------------------

    def select_frame_by_name(self, name):
        # set button color for selected button
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.utilitys_button.configure(fg_color=("gray75", "gray25") if name == "utilitys" else "transparent")
        self.members_button.configure(fg_color=("gray75", "gray25") if name == "members" else "transparent")

        # show selected frame
        if name == "home":
            self.home_frame.tkraise()
        if name == "utilitys":
            self.utilitys_frame.tkraise()
        if name == "members":
            self.members_frame.tkraise()
        if name == "login":
            self.login_frame.tkraise()

    def home_button_event(self):
        self.select_frame_by_name("home")

    def utilitys_button_event(self):
        self.select_frame_by_name("utilitys")

    def members_button_event(self):
        self.select_frame_by_name("members")
    
    def login_button_event(self, name):
        self.select_frame_by_name("login") 
    
    def register_window(self):
        self.register_frame.tkraise()
    


    # ----------Get Values From Entries-------------
    def getvalue(self):
        self.name2 = self.name1.get()
        self.apellido2 = self.apellido1.get()
        self.cedula2 = self.cedula.get()
        self.edad2 = self.edad.get()
        print(type(self.cedula2 and self.edad2))
        # -----------------Warnings-----------------
        if not self.name2 or not self.apellido2 or not self.edad2 or not self.cedula2:
            CTkMessagebox(self, title="Error",
                          message="Please put your Nombre, Apellido, Cedula and Edad",
                          icon="cancel",
                          sound=True,
                          fade_in_duration=1,
                          width=200,
                          height=50)
        elif len(self.name2 and self.apellido2) >= 25:
            CTkMessagebox(self, title="Error",
                          message="Please the Nombre and Apellido must be 16 characters long",
                          icon="cancel",
                          sound=True,
                          fade_in_duration=1,
                          width=200,
                          height=50)
        elif len(self.cedula2) != 11:
            CTkMessagebox(self, title="Error",
                          message="Please enter the 11 digits cedula",
                          icon="cancel",
                          sound=True,
                          fade_in_duration=1,
                          width=200,
                          height=50)
        #elif (self.cedula2 and self.edad2) != int:
            #CTkMessagebox(self, title="Error",
                          #message="Cedula and Edad must numbers only",
                          #icon="cancel",
                          #sound=True,
                          #fade_in_duration=1,
                          #width=200,
                          #height=50)
        else:
            # ----------Load Data into Excel Sheet------------------
            self.wkb = load_workbook("ProjetoDoClube/Data1.xlsx")
            self.wkbs = self.wkb.active
            self.row_values = [self.name2, self.apellido2,self.edad2, self.cedula2]
            self.wkbs.append(self.row_values)
            self.wkb.save("ProjetoDoClube/Data1.xlsx")
            # Insert Data in Treeview
            self.treevie.insert("", "end", values=self.row_values)
            # Delete Input in Entries
            self.name1.delete("0", "end")
            self.apellido1.delete("0", "end")
            self.cedula.delete("0", "end")

    def load_data(self,):
        # ----------------Load the XLSX file---------------------------
        self.wb = load_workbook("C:/Users/anavi/Downloads/Andre Programming/ProjetoDoClube/Data1.xlsx")
        self.ws = self.wb.active
        self.list_values = list(self.ws.values)
        for self.col_name in self.list_values[0]:
            self.treevie.heading(self.col_name, text=self.col_name)

        for self.list_tuples in self.list_values[1:]:
            self.treevie.insert('', 'end', values=self.list_tuples)
            # print(self.list_values[1:])
        self.submit_button2.configure(state="disabled")
        print(self.submit_button2.cget("state"))

    def check_login(self):
        username = self.lo_username_entry.get()
        password = self.lo_password_entry.get()
        response = requests.post("http://127.0.0.1:5000/login", json={"username": username, "password": password})
        token = response.json()
        print(token)
        # Simple hard-coded login credentials
        if response.status_code == 200:
            self.login_frame.destroy()
            self.register_frame.destroy()
            
        else:
            CTkMessagebox(self, title="Login Failed", message="Invalid username or password", icon="cancel")
    
    
    
    def check_register(self):
        username = self.re_username_entry.get()
        password = self.re_password_entry.get()
        register = requests.post("http://127.0.0.1:5000/register_ussers", json={"username": username, "password": password})
        print(register)
        if register.status_code == 201:
            self.register_label = customtkinter.CTkLabel(master=self.register_frame, text="User registered successfully", text_color="black", font=("Arial", 20))
            self.register_label.pack()
app = MainApp()
app.mainloop()